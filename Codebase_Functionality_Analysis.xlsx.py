from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule

wb = Workbook()

# ============ SHEET 1: Summary ============
ws = wb.active
ws.title = "Summary"

thin = Side(border_style="thin", color="D0D5DD")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# Title
ws["A1"] = "Clinic Management System — Codebase Functionality Audit"
ws["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
ws["A1"].fill = PatternFill("solid", start_color="2E37A4")
ws["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.row_dimensions[1].height = 32
ws.merge_cells("A1:E1")

ws["A2"] = "Project: clinic-management-system (Next.js 16.2.4, React 19, TailwindCSS v4)"
ws["A2"].font = Font(name="Arial", size=10, italic=True, color="475467")
ws.merge_cells("A2:E2")

ws["A3"] = "Date: 2026-05-13"
ws["A3"].font = Font(name="Arial", size=10, italic=True, color="475467")
ws.merge_cells("A3:E3")

# Headers
headers = ["#", "Module / Feature", "Status / What's Implemented", "What's Missing", "Completion %"]
for col, h in enumerate(headers, start=1):
    c = ws.cell(row=5, column=col, value=h)
    c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="2E37A4")
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = border
ws.row_dimensions[5].height = 30

rows = [
    (1, "Login (app/page.tsx)",
        "Email/password UI, show/hide toggle, button enables when both fields filled.",
        "handleLogin only does router.push to /admin/patients. No real auth, no validation, no session, no JWT/cookie, no middleware.",
        15),
    (2, "Forgot Password / OTP / Reset",
        "5-step flow in forgot-password/page.tsx with 30s OTP timer, password match check.",
        "auth/login/page.tsx returns null. auth/otp/page.tsx and auth/new-password/page.tsx are 0-byte empty files. No email send, no OTP verify, no password update API.",
        25),
    (3, "Dashboard / Overview",
        "4 stat cards, upcoming appointments table, 3 performance cards.",
        "All data hardcoded. No charts, no date filters, no real KPIs. 'View All' and 'Take Action' buttons non-functional.",
        20),
    (4, "Sidebar / Header / Layout",
        "7 main + 2 bottom nav items, active-route highlight via usePathname.",
        "Sidebar collapse button non-functional. Header search/dark-mode/bell/profile all decorative. Hardcoded user 'Amit Singh / Patient'.",
        35),
    (5, "Patient List",
        "Table view, search input UI, filter button, count badge, 'Add New Patient' link.",
        "8 hardcoded patients (all named 'Dr. ...'). No fetch/pagination/sort. Search input has no onChange. View/Edit/More buttons non-functional. Label says '29 patients' but shows 8.",
        15),
    (6, "Patient Add — 6-tab Stepper (3,043 lines)",
        "Massive stepper: RMO Consultation, Patient Details, Main Consultation, Infusion/Rehab/Aesthetics, Test, Summary. Tab/sub-section navigation works.",
        "Every input is uncontrolled. No form state, no validation, no save. 'Add Patient' button has no onClick. Some chars render as '??'.",
        20),
    ("6a", "↳ RMO Consultation (6 sub-sections)",
        "Informant, Demographics, Medical History, Social History, Personal History, Examination Summary. Accordions toggle correctly.",
        "Thorough static form fields; zero data wiring.",
        25),
    ("6b", "↳ Patient Details (Blood Group)",
        "6 Rh-factor radio questions.",
        "Very thin — name suggests far broader scope (demographics, contact, ID).",
        10),
    ("6c", "↳ Main Consultation",
        "Custom calendar widget + 8 time slots, local selection works.",
        "No clinical fields (no chief complaints, history of present illness, diagnosis, prescription, vitals).",
        20),
    ("6d", "↳ Infusion, Rehab & Aesthetics",
        "3 metric cards + recent activity list.",
        "All numbers hardcoded; no scheduling/treatment data entry. Doesn't match tab name.",
        10),
    ("6e", "↳ Test (10 lab panels)",
        "Routine, Male/Female Hormonal, Dutch, Autoimmune, Genetic, Tumor Markers, Micronutrient, Metabolic, Stool — large static lab-value input grid.",
        "All inputs uncontrolled. No reference ranges, no upload of lab reports, no historical comparison, no save.",
        25),
    ("6f", "↳ Summary",
        "6 stat cards + 7-row table with filter dropdowns and pagination buttons.",
        "Looks like a patient-pool dashboard rather than an individual summary. Filters, pagination, 'Add' button all non-functional.",
        15),
    (7, "Appointments List",
        "Table with type/status badges, 'All Status' & 'All Dates' filter selects.",
        "10 hardcoded rows (all 'Sumit Mittal'). Filter/search non-functional. Column header bug: 'Appointments.Status'. Label '35 appointments' doesn't match 10 rows.",
        15),
    (8, "Appointments Add",
        "4-step wizard: Patient Info → Appointment Details → Additional Info → Review. Step nav, Previous/Next/Book buttons.",
        "Every field uncontrolled. Review step shows 'Not provided'/'Not selected' hardcoded. 'Book Appointment' does nothing. Unreachable default switch case.",
        20),
    (9, "Staff List",
        "Table with avatar, role, status badges, search input.",
        "6 hardcoded staff (all 'Sumit Mittal' / 'sumit1@gmail.com'). Search non-functional. Details/Edit/More buttons non-functional.",
        15),
    (10, "Staff Add",
        "Static form: Basic Info, Professional Info, Qualifications, Bio, Account Info.",
        "Zero useState — fully uncontrolled. No file/photo upload. No validation. 'Add Staff Member' buttons non-functional.",
        15),
    (11, "Departments",
        "6 polished hardcoded department cards (Admin, Reception, RMO, Infusion, Rehab, Aesthetics) with theme colors and metric trios.",
        "No CRUD. No real link between departments and staff/patients. 'View Details' non-functional.",
        15),
    (12, "Invoices List",
        "13 hardcoded invoices, status badges, status filter dropdown opens.",
        "No 'New Invoice' button. Search non-functional. Filter doesn't actually filter. More-options non-functional.",
        20),
    (13, "Invoice Detail ([id])",
        "Dynamic route reads params.id and displays it. Items table, financial summary, payment history.",
        "Shows identical hardcoded content regardless of which invoice ID is in URL. Print/Edit buttons non-functional. No fetch by ID.",
        15),
    (14, "Reports page",
        "Sidebar entry exists.",
        "File reports/page.tsx is completely empty (0 lines). Would render nothing or error.",
        0),
    (15, "Settings",
        "Sidebar link to /admin/settings.",
        "Route directory does not exist — would 404.",
        0),
    (16, "Help & Support",
        "Sidebar link to /admin/help.",
        "Route directory does not exist — would 404.",
        0),
    (17, "Backend / API / Database",
        "—",
        "Zero API routes. Zero fetch() calls. No DB/ORM. No env vars. No auth provider. Default empty next.config.ts.",
        0),
    (18, "Global Search",
        "Search inputs in header and on every list page.",
        "No onChange handler anywhere. No filter/search logic.",
        5),
    (19, "Notifications",
        "Bell icon with red-dot badge.",
        "No dropdown, no feed, no state, no API.",
        5),
    (20, "Profile Management",
        "Profile button in header (avatar + name + chevron).",
        "Hardcoded 'Amit Singh / Patient'. No dropdown, no profile page, no edit, no logout.",
        5),
    (21, "Dark Mode",
        "Moon icon button in header.",
        "No state, no next-themes, no dark tokens applied. Whole UI uses hardcoded hex colors.",
        5),
    (22, "Form Validation",
        "Login (both fields non-empty), reset password (match check).",
        "No Zod / react-hook-form / Yup. Required asterisks visual only.",
        5),
    (23, "State Management / Data Layer",
        "useState in 5 files (local UI state only).",
        "No Redux/Zustand/Context/React Query/SWR. No shared types directory. lib/utils.ts only exports cn().",
        0),
]

row_idx = 6
for r in rows:
    for col_idx, val in enumerate(r, start=1):
        c = ws.cell(row=row_idx, column=col_idx, value=val)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        if col_idx == 1:
            c.alignment = Alignment(horizontal="center", vertical="top")
            c.font = Font(name="Arial", size=10, bold=True)
        if col_idx == 2:
            c.font = Font(name="Arial", size=10, bold=True, color="101828")
        if col_idx == 5:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name="Arial", size=11, bold=True)
            c.number_format = '0"%"'
    row_idx += 1

# Totals row with formula (weighted average — simple average of leaf items)
# Use average of completion % column from data rows, excluding the parent row 6 (parent of 6a-6f)
last_data_row = row_idx - 1
totals_row = row_idx
ws.cell(row=totals_row, column=1, value="").border = border
ws.cell(row=totals_row, column=2, value="OVERALL COMPLETION (avg of leaf features)")
ws.cell(row=totals_row, column=3, value="")
ws.cell(row=totals_row, column=4, value="")
ws.cell(row=totals_row, column=5, value=f"=AVERAGE(E6:E{last_data_row})")
for col_idx in range(1, 6):
    c = ws.cell(row=totals_row, column=col_idx)
    c.fill = PatternFill("solid", start_color="F4F5FF")
    c.font = Font(name="Arial", size=11, bold=True, color="2E37A4")
    c.border = border
    if col_idx == 5:
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.number_format = '0.0"%"'
    elif col_idx == 2:
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    else:
        c.alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[totals_row].height = 26

# Column widths
ws.column_dimensions["A"].width = 6
ws.column_dimensions["B"].width = 36
ws.column_dimensions["C"].width = 55
ws.column_dimensions["D"].width = 55
ws.column_dimensions["E"].width = 14

# Row heights for wrapping
for r in range(6, last_data_row + 1):
    ws.row_dimensions[r].height = 48

# Conditional formatting on completion % column (red → yellow → green)
rule = ColorScaleRule(
    start_type="num", start_value=0, start_color="F97066",
    mid_type="num", mid_value=50, mid_color="FEC84B",
    end_type="num", end_value=100, end_color="32D583",
)
ws.conditional_formatting.add(f"E6:E{last_data_row}", rule)

# Freeze header
ws.freeze_panes = "A6"

# ============ SHEET 2: Gaps & Priorities ============
ws2 = wb.create_sheet("Biggest Gaps")
ws2["A1"] = "Biggest Gaps — Priority Order"
ws2["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
ws2["A1"].fill = PatternFill("solid", start_color="B42318")
ws2["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws2.merge_cells("A1:C1")
ws2.row_dimensions[1].height = 28

gap_headers = ["Priority", "Gap", "Why it matters"]
for col, h in enumerate(gap_headers, start=1):
    c = ws2.cell(row=3, column=col, value=h)
    c.font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill("solid", start_color="2E37A4")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = border
ws2.row_dimensions[3].height = 26

gaps = [
    (1, "Backend / API / Database layer", "0% — foundational blocker for every other feature."),
    (2, "Authentication & session", "Login bypass is router.push. No protected routes, no JWT/cookie, no role-based access."),
    (3, "Patient persistence", "The largest UI feature (3,043-line form) has no way to save anything."),
    (4, "Reports, Settings, Help", "Reports file is empty (0 lines). Settings and Help routes don't exist (would 404)."),
    (5, "CRUD operations", "No Create/Update/Delete works on any list (patients, appointments, staff, invoices, departments)."),
    (6, "Form state & validation", "Every data-entry form is fully uncontrolled. No Zod/react-hook-form."),
    (7, "Dynamic detail routes", "Only invoice has [id] route — and it ignores the ID."),
    (8, "Search, filtering, pagination", "Search inputs everywhere; none have onChange. Filters are decorative."),
    (9, "Profile management & dark mode", "Header chrome (profile dropdown, dark-mode toggle, bell) is entirely decorative."),
]
for i, g in enumerate(gaps, start=4):
    for col_idx, val in enumerate(g, start=1):
        c = ws2.cell(row=i, column=col_idx, value=val)
        c.font = Font(name="Arial", size=10)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        c.border = border
        if col_idx == 1:
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.font = Font(name="Arial", size=11, bold=True, color="B42318")
        if col_idx == 2:
            c.font = Font(name="Arial", size=10, bold=True)
    ws2.row_dimensions[i].height = 36

ws2.column_dimensions["A"].width = 10
ws2.column_dimensions["B"].width = 38
ws2.column_dimensions["C"].width = 70
ws2.freeze_panes = "A4"

# ============ SHEET 3: Stats ============
ws3 = wb.create_sheet("Stats")
ws3["A1"] = "Quick Stats"
ws3["A1"].font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
ws3["A1"].fill = PatternFill("solid", start_color="2E37A4")
ws3["A1"].alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws3.merge_cells("A1:B1")
ws3.row_dimensions[1].height = 28

stats_data = [
    ("Stack", "Next.js 16.2.4 / React 19 / TailwindCSS v4"),
    ("Total pages (app/**)", "20"),
    ("Empty (0-byte) page files", "3  (auth/login, auth/otp, auth/new-password, reports)"),
    ("API routes", "0"),
    ("fetch() calls in code", "0"),
    ("Files using useState", "5"),
    ("Form validation library", "None (no Zod / react-hook-form / Yup)"),
    ("Global state library", "None"),
    ("DB / ORM", "None"),
    ("Largest file", "app/admin/(dashboard)/patients/add/page.tsx — 3,043 lines"),
    ("Tests / CI", "None"),
    ("Total git commits", "5"),
    ("Overall completion estimate", "~15–18% of a production-ready system"),
]
for i, (k, v) in enumerate(stats_data, start=3):
    a = ws3.cell(row=i, column=1, value=k)
    b = ws3.cell(row=i, column=2, value=v)
    a.font = Font(name="Arial", size=10, bold=True, color="101828")
    b.font = Font(name="Arial", size=10, color="344054")
    a.fill = PatternFill("solid", start_color="F9FAFB")
    a.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    b.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)
    a.border = border
    b.border = border
    ws3.row_dimensions[i].height = 24

ws3.column_dimensions["A"].width = 32
ws3.column_dimensions["B"].width = 70

out_path = "/sessions/ecstatic-vigilant-noether/mnt/clinic-management-system/Codebase_Functionality_Analysis.xlsx"
wb.save(out_path)
print(f"saved: {out_path}")
